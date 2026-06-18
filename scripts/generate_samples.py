"""
scripts/generate_samples.py
===========================
Create realistic sample logistics Excel datasets for testing the copilot.

Generates three scenario files in ``data/samples/``:
* sample_baseline.xlsx     – mostly nominal operations
* sample_disruption.xlsx   – delay + fuel-spike + staffing crisis
* sample_mixed.xlsx        – blended, multi-route portfolio

Run:  python scripts/generate_samples.py
"""
from __future__ import annotations

import sys
from pathlib import Path

import numpy as np
import pandas as pd

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
from config import SAMPLE_DIR  # noqa: E402

RNG = np.random.default_rng(42)

ROUTES = [
    "Shanghai-LA", "Rotterdam-NYC", "Singapore-Mumbai",
    "Dubai-Hamburg", "Mumbai-Dubai", "LA-Tokyo",
]
SIGNALS = [
    "port congestion", "fuel price surge", "labour strike warning",
    "storm forecast", "customs delay", "carrier capacity tight",
    "", "", "",  # many shipments have no signal
]


def _make(n: int, *, delay_mu: float, delay_sd: float, fuel_mu: float,
          fuel_sd: float, staff_mu: float, complaint_lam: float,
          spike_frac: float = 0.0, prefix: str = "S") -> pd.DataFrame:
    dates = pd.date_range("2025-01-01", periods=n, freq="D")
    delay = np.clip(RNG.normal(delay_mu, delay_sd, n), 0, None)
    fuel = np.clip(RNG.normal(fuel_mu, fuel_sd, n), 500, None)
    staff = np.clip(RNG.normal(staff_mu, 3, n).round(), 2, None).astype(int)
    complaints = RNG.poisson(complaint_lam, n)

    # Inject extreme spikes into a fraction of rows.
    if spike_frac:
        k = max(1, int(n * spike_frac))
        idx = RNG.choice(n, k, replace=False)
        delay[idx] *= RNG.uniform(3, 6, k)
        fuel[idx] *= RNG.uniform(2, 3, k)
        staff[idx] = np.clip(staff[idx] - RNG.integers(5, 12, k), 1, None)
        complaints[idx] += RNG.integers(5, 20, k)

    return pd.DataFrame({
        "ShipmentID": [f"{prefix}{i:04d}" for i in range(1, n + 1)],
        "Route": RNG.choice(ROUTES, n),
        "Date": dates,
        "Delay": delay.round(1),
        "FuelCost": fuel.round(0),
        "StaffCount": staff,
        "Complaints": complaints,
        "ExternalSignals": RNG.choice(SIGNALS, n),
    })


def _save(df: pd.DataFrame, name: str) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils.dataframe import dataframe_to_rows

    path = SAMPLE_DIR / name
    wb = Workbook()
    ws = wb.active
    ws.title = "logistics"
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    # Professional font + bold header (per xlsx conventions).
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True)
    for r in ws.iter_rows(min_row=2):
        for c in r:
            c.font = Font(name="Arial")
    for col_cells in ws.columns:
        width = max(len(str(c.value)) if c.value is not None else 0
                    for c in col_cells) + 2
        ws.column_dimensions[col_cells[0].column_letter].width = min(width, 24)
    wb.save(path)
    return path


def main() -> None:
    baseline = _make(60, delay_mu=3, delay_sd=2, fuel_mu=4000, fuel_sd=400,
                     staff_mu=18, complaint_lam=0.3, prefix="B")
    disruption = _make(60, delay_mu=10, delay_sd=6, fuel_mu=5200, fuel_sd=900,
                       staff_mu=12, complaint_lam=2.0, spike_frac=0.25, prefix="D")
    mixed = _make(120, delay_mu=6, delay_sd=5, fuel_mu=4600, fuel_sd=700,
                  staff_mu=15, complaint_lam=1.0, spike_frac=0.12, prefix="M")

    for df, name in [(baseline, "sample_baseline.xlsx"),
                     (disruption, "sample_disruption.xlsx"),
                     (mixed, "sample_mixed.xlsx")]:
        p = _save(df, name)
        print(f"wrote {p}  ({len(df)} rows)")


if __name__ == "__main__":
    main()
